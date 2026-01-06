"""
Export views for exporting expense data to CSV, XLSX, and PDF formats.
"""
import csv
import io
from datetime import datetime, date
from decimal import Decimal

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Expense, Category

from apps.common.utils import error_response


class ExportExpensesView(APIView):
    """
    API view to export expenses data in CSV, XLSX, or PDF format.
    
    Query Parameters:
    - export_format: Export format ('csv', 'xlsx', 'pdf'). Required. Default: 'csv'
    - start_date: Filter expenses from this date (YYYY-MM-DD). Optional.
    - end_date: Filter expenses until this date (YYYY-MM-DD). Optional. Defaults to current date if start_date is provided.
    - category: Filter by category ID. Optional.
    - category: Filter by category ID. Optional.
    
    Notes:
    - Only the export_format is required, all other fields are optional
    - If no date filters are provided, exports all expenses
    - If start_date is provided without end_date, end_date defaults to current date
    - Requires authentication
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Check authentication
            if not request.user.is_authenticated:
                return Response(
                    {
                        'success': False,
                        'error': 'Authentication required',
                        'message': 'Please log in to export expenses'
                    },
                    status=status.HTTP_401_UNAUTHORIZED
                )

            # Get query parameters
            export_format = request.query_params.get('export_format', 'csv').lower()
            start_date = request.query_params.get('start_date', '').strip()
            end_date = request.query_params.get('end_date', '').strip()
            category_id = request.query_params.get('category', '').strip()
            category_id = request.query_params.get('category', '').strip()

            # Validate format (only required field)
            if export_format not in ['csv', 'xlsx', 'pdf']:
                return Response(
                    {
                        'success': False,
                        'error': 'Invalid format',
                        'message': 'Invalid export_format. Supported formats: csv, xlsx, pdf'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Build queryset with filters
            queryset = Expense.objects.filter(user=request.user).select_related('category', 'user')
            
            # Track applied filters for response metadata
            filters_applied = []

            # Date filtering with smart defaults
            start = None
            end = None
            
            if start_date:
                try:
                    start = datetime.strptime(start_date, '%Y-%m-%d').date()
                    queryset = queryset.filter(date__gte=start)
                    filters_applied.append(f"Start Date: {start_date}")
                except ValueError:
                    return Response(
                        {
                            'success': False,
                            'error': 'Invalid date format',
                            'message': 'Invalid start_date format. Please use YYYY-MM-DD (e.g., 2025-01-01)'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # If start_date is provided but end_date is not, default end_date to current date
            if start_date and not end_date:
                end = date.today()
                queryset = queryset.filter(date__lte=end)
                filters_applied.append(f"End Date: {end.strftime('%Y-%m-%d')} (today)")
            elif end_date:
                try:
                    end = datetime.strptime(end_date, '%Y-%m-%d').date()
                    queryset = queryset.filter(date__lte=end)
                    filters_applied.append(f"End Date: {end_date}")
                except ValueError:
                    return Response(
                        {
                            'success': False,
                            'error': 'Invalid date format',
                            'message': 'Invalid end_date format. Please use YYYY-MM-DD (e.g., 2025-12-31)'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Validate date range
            if start and end and start > end:
                return Response(
                    {
                        'success': False,
                        'error': 'Invalid date range',
                        'message': 'Start date cannot be after end date'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            if category_id:
                try:
                    cat_id = int(category_id)
                    queryset = queryset.filter(category_id=cat_id)
                    # Get category name for display
                    try:
                        category = Category.objects.get(id=cat_id, user=request.user)
                        filters_applied.append(f"Category: {category.name}")
                    except Category.DoesNotExist:
                        filters_applied.append(f"Category ID: {cat_id}")
                except ValueError:
                    return Response(
                        {
                            'success': False,
                            'error': 'Invalid category',
                            'message': 'Category ID must be a valid number'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Order by date descending
            queryset = queryset.order_by('-date')
            
            # Generate export based on format
            expenses = list(queryset)
            
            if export_format == 'csv':
                return self._export_csv(expenses, filters_applied)
            elif export_format == 'xlsx':
                return self._export_xlsx(expenses, filters_applied)
            elif export_format == 'pdf':
                return self._export_pdf(expenses, filters_applied)
                
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'error': 'Export failed',
                    'message': f'An unexpected error occurred while exporting: {str(e)}'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_expense_data(self, expenses):
        """Convert expense objects to list of dictionaries for export."""
        data = []
        for expense in expenses:
            data.append({
                'ID': expense.id,
                'Date': expense.date.strftime('%Y-%m-%d'),
                'Title': expense.title,
                'Description': expense.description or '',
                'Category': expense.category.name if expense.category else 'Uncategorized',
                'Amount': float(expense.amount),
                'Recurring': 'Yes' if expense.is_recurring else 'No',
                'Frequency': expense.recurring_frequency or 'N/A',
            })
        return data

    def _export_csv(self, expenses, filters_applied=None):
        """Export expenses to CSV format."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="expenses_{timestamp}.csv"'
        # Add CORS header for download
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'

        data = self._get_expense_data(expenses)
        
        writer = csv.writer(response)
        
        # Add header info
        writer.writerow(['SpendWise - Expense Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        if filters_applied:
            writer.writerow([f'Filters: {", ".join(filters_applied)}'])
        else:
            writer.writerow(['Filters: All expenses (no filters applied)'])
        writer.writerow([f'Total Records: {len(expenses)}'])
        writer.writerow([])  # Empty row
        
        if not data:
            writer.writerow(['No expenses found matching your criteria'])
            return response

        # Write data
        field_writer = csv.DictWriter(response, fieldnames=data[0].keys())
        field_writer.writeheader()
        field_writer.writerows(data)
        
        # Add summary
        total_amount = sum(float(row['Amount']) for row in data)
        writer.writerow([])
        writer.writerow(['Summary'])
        writer.writerow([f'Total Amount: PHP {total_amount:.2f}'])
        writer.writerow([f'Number of Expenses: {len(data)}'])
        if data:
            avg_amount = total_amount / len(data)
            writer.writerow([f'Average Amount: PHP {avg_amount:.2f}'])

        return response

    def _export_xlsx(self, expenses, filters_applied=None):
        """Export expenses to XLSX (Excel) format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        data = self._get_expense_data(expenses)

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        title_font = Font(bold=True, size=14)
        info_font = Font(italic=True, size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add report header info
        ws['A1'] = 'SpendWise - Expense Report'  
        ws['A1'].font = title_font
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = info_font
        if filters_applied:
            ws['A3'] = f'Filters: {", ".join(filters_applied)}'
        else:
            ws['A3'] = 'Filters: All expenses (no filters applied)'
        ws['A3'].font = info_font
        ws['A4'] = f'Total Records: {len(expenses)}'
        ws['A4'].font = info_font
        
        data_start_row = 6  # Start data after header info

        if not data:
            ws.cell(row=data_start_row, column=1, value='No expenses found matching your criteria')
        else:
            # Write headers
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=data_start_row, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data
            for row_num, row_data in enumerate(data, data_start_row + 1):
                for col_num, (key, value) in enumerate(row_data.items(), 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    if key == 'Amount':
                        cell.number_format = 'PHP #,##0.00'

            # Adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(header)
                for row in range(data_start_row + 1, len(data) + data_start_row + 1):
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Add summary section
            summary_row = len(data) + data_start_row + 2
            ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
            
            total_amount = sum(expense.amount for expense in expenses)
            ws.cell(row=summary_row + 1, column=1, value="Total Expenses:")
            ws.cell(row=summary_row + 1, column=2, value=float(total_amount))
            ws.cell(row=summary_row + 1, column=2).number_format = 'PHP #,##0.00'
            
            ws.cell(row=summary_row + 2, column=1, value="Number of Expenses:")
            ws.cell(row=summary_row + 2, column=2, value=len(expenses))
            
            if expenses:
                avg_amount = float(total_amount) / len(expenses)
                ws.cell(row=summary_row + 3, column=1, value="Average Amount:")
                ws.cell(row=summary_row + 3, column=2, value=avg_amount)
                ws.cell(row=summary_row + 3, column=2).number_format = 'PHP #,##0.00'

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="expenses_{timestamp}.xlsx"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'

        return response

    def _export_pdf(self, expenses, filters_applied=None):
        """Export expenses to PDF format - Professional and presentable layout."""
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )

        elements = []
        styles = getSampleStyleSheet()

        # ========== CUSTOM STYLES ==========
        # Main title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=5,
            alignment=1  # Center
        )
        
        # Subtitle
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#64748b'),
            alignment=1,
            spaceAfter=3
        )
        
        # Section headers
        section_header_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#334155'),
            spaceBefore=15,
            spaceAfter=8
        )
        
        # Breakdown section header - tighter spacing
        breakdown_header_style = ParagraphStyle(
            'BreakdownHeader',
            parent=styles['Heading2'],
            fontSize=14,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#334155'),
            spaceBefore=12,
            spaceAfter=0  # No spacing after header - very compact
        )
        
        # Info text
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#94a3b8'),
            alignment=1
        )

        # ========== HEADER SECTION ==========
        # Title with decorative line
        elements.append(Paragraph("SpendWise - Expense Report", title_style))
        elements.append(Paragraph("Financial Summary & Transaction Details", subtitle_style))
        elements.append(Spacer(1, 5))
        
        # Decorative line
        line_table = Table([['']], colWidths=[9*inch])
        line_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor('#3b82f6')),
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 10))

        # ========== REPORT INFO BOX ==========
        generated_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        filters_text = ', '.join(filters_applied) if filters_applied else 'No filters applied (showing all expenses)'
        
        info_data = [
            ['Generated:', generated_date, 'Total Records:', str(len(expenses))],
            ['Filters:', filters_text, '', ''],
        ]
        info_table = Table(info_data, colWidths=[1.2*inch, 3.5*inch, 1.3*inch, 1.5*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#475569')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('SPAN', (1, 1), (3, 1)),  # Span filters across columns
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 15))

        data = self._get_expense_data(expenses)

        if not data:
            # ========== NO DATA MESSAGE ==========
            no_data_style = ParagraphStyle(
                'NoData',
                parent=styles['Normal'],
                fontSize=14,
                alignment=1,
                textColor=colors.HexColor('#94a3b8'),
                spaceBefore=30
            )
            elements.append(Paragraph("No expenses found matching your criteria", no_data_style))
            elements.append(Paragraph("Try adjusting your filters or date range", info_style))
        else:
            # ========== SUMMARY CARDS SECTION ==========
            elements.append(Paragraph("Financial Summary", section_header_style))
            
            total_amount = sum(expense.amount for expense in expenses)
            avg_amount = float(total_amount / len(expenses)) if expenses else 0.0
            
            # Calculate category breakdown
            category_totals = {}
            for expense in expenses:
                cat_name = expense.category.name if expense.category else 'Uncategorized'
                category_totals[cat_name] = category_totals.get(cat_name, 0) + float(expense.amount)
            
            # Top category
            top_category = max(category_totals.items(), key=lambda x: x[1]) if category_totals else ('N/A', 0)
            
            # Summary cards data
            summary_cards = [
                ['TOTAL SPENT', 'TRANSACTIONS', 'AVERAGE', 'TOP CATEGORY'],
                [f'PHP {float(total_amount):,.2f}', str(len(expenses)), f'PHP {avg_amount:,.2f}', f'{top_category[0]}'],
                ['Total expenses', 'Number of items', 'Per transaction', f'PHP {top_category[1]:,.2f}'],
            ]
            
            summary_cards_table = Table(summary_cards, colWidths=[2.2*inch, 2.2*inch, 2.2*inch, 2.2*inch])
            summary_cards_table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                # Value row
                ('BACKGROUND', (0, 1), (-1, 1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1e293b')),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, 1), 16),
                ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
                # Description row
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor('#64748b')),
                ('FONTSIZE', (0, 2), (-1, 2), 8),
                ('ALIGN', (0, 2), (-1, 2), 'CENTER'),
                # Borders and padding - more generous padding
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(summary_cards_table)
            elements.append(Spacer(1, 3))

            # ========== BREAKDOWN SECTION ==========
            # Category Breakdown
            elements.append(Paragraph("Category Breakdown", breakdown_header_style))
            
            # Category breakdown (top 5)
            sorted_categories = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:5]
            category_data = [['Category', 'Amount', '% of Total']]
            for cat_name, amount in sorted_categories:
                pct = (amount / float(total_amount)) * 100 if total_amount else 0
                # Truncate category name if too long
                cat_display = cat_name[:15] + '...' if len(cat_name) > 15 else cat_name
                category_data.append([cat_display, f'PHP {amount:,.2f}', f'{pct:.1f}%'])
            
            category_table = Table(category_data, colWidths=[3.5*inch, 2.5*inch, 1.5*inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f3ff')]),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#ddd6fe')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd6fe')),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(category_table)
            elements.append(Spacer(1, 20))

            # ========== DETAILED TRANSACTIONS TABLE ==========
            elements.append(Paragraph("Detailed Transactions", section_header_style))
            
            # Prepare table data with better formatting - include description
            headers = ['#', 'Date', 'Title', 'Description', 'Category', 'Amount', 'Recurring']
            table_data = [headers]
            
            for idx, row in enumerate(data, 1):
                # Truncate title, description, and category to prevent table breaking
                title_display = row['Title'][:18] + '...' if len(row['Title']) > 18 else row['Title']
                desc_display = row['Description'][:30] + '...' if len(row['Description']) > 30 else row['Description']
                if not desc_display:
                    desc_display = '-'
                
                # Truncate category name if too long
                category_display = row['Category'][:15] + '...' if len(row['Category']) > 15 else row['Category']
                

                table_data.append([
                    str(idx),
                    row['Date'],
                    title_display,
                    desc_display,
                    category_display,
                    f"PHP {row['Amount']:,.2f}",
                    row['Recurring']
                ])

            # Create data table with professional styling - adjusted column widths for better spacing
            col_widths = [0.35*inch, 0.85*inch, 1.4*inch, 2.0*inch, 1.5*inch, 1.0*inch, 0.75*inch]
            table = Table(table_data, colWidths=col_widths)
            
            table.setStyle(TableStyle([
                # Header styling - more padding for headers
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
                ('TOPPADDING', (0, 0), (-1, 0), 15),
                ('LEFTPADDING', (0, 0), (-1, 0), 10),
                ('RIGHTPADDING', (0, 0), (-1, 0), 10),
                # Data rows
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                # Alignment
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # # column
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Date column
                ('ALIGN', (2, 0), (2, -1), 'LEFT'),    # Title column
                ('ALIGN', (3, 0), (3, -1), 'LEFT'),    # Description column
                ('ALIGN', (4, 0), (4, -1), 'LEFT'),    # Category column
                ('ALIGN', (5, 0), (5, -1), 'RIGHT'),   # Amount column
                ('ALIGN', (6, 0), (6, -1), 'CENTER'),  # Recurring column
                # Borders
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#3b82f6')),
                ('INNERGRID', (0, 1), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                # More padding for data rows
                ('TOPPADDING', (0, 1), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
                ('LEFTPADDING', (0, 1), (-1, -1), 10),
                ('RIGHTPADDING', (0, 1), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

            # ========== FOOTER ==========
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#94a3b8'),
                alignment=1
            )
            elements.append(Paragraph("─" * 80, footer_style))
            elements.append(Paragraph(
                f"This report was automatically generated by SpendWise - {datetime.now().strftime('%Y')}",
                footer_style
            ))
            elements.append(Paragraph(
                "For questions or support, please contact your administrator",
                footer_style
            ))

        # Build PDF
        doc.build(elements)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="expense_report_{timestamp}.pdf"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'

        return response
